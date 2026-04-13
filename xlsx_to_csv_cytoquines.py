#!/usr/bin/env python3

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import csv


def _extract_tables(sheet, table):

    start_row, end_row = 8, 73
    start_col, end_col = 1, 8

    for row in sheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_col,
        max_col=end_col,
        values_only=True,
    ):
        table.append(list(row))

    return table


def _clean_dataframe(data):

    data = data.iloc[:, 0:]
    data = data[data.iloc[:, 0] != "Type"]
    new_columns = ["Sample", "Location"]

    for col in data.columns[2:]:
        new_columns.append(col.split("(")[0].strip())

    data.columns = new_columns

    for col in data.columns[2:]:
        # Replace *** by NaN
        data[col] = data[col].replace("***", np.nan)
        # Fill blank spaces with 0
        data[col] = data[col].replace("", "0")
        data[col] = pd.to_numeric(data[col].str.replace(",", "."), errors="coerce")

    return data


if __name__ == "__main__":
    wb = load_workbook("data/citoquinas_3.xlsx")
    results = wb["FI-Bckg"]
    counts = wb["Bead Count"]
    obs_conc = wb["Obs Conc"]

    results_table = []
    count_table = []
    obs_conc_table = []

    results = _extract_tables(results, results_table)
    counts = _extract_tables(counts, count_table)
    obs_conc = _extract_tables(obs_conc, obs_conc_table)

    df_results = pd.DataFrame(results[1:], columns=results[0])
    df_counts = pd.DataFrame(counts[1:], columns=counts[0])
    df_obs_conc = pd.DataFrame(obs_conc[1:], columns=obs_conc[0])

    df_results = _clean_dataframe(df_results)
    df_counts = _clean_dataframe(df_counts)
    df_obs_conc = _clean_dataframe(df_obs_conc)

    with open("data/citokines_3.csv", "w", newline="") as f:
        #  Count
        f.write('"DataType:","Count"\n')
        df_counts.to_csv(f, index=False, quoting=csv.QUOTE_ALL)

        # Empty line
        f.write("\n")
        # Result
        f.write('"DataType:","Result"\n')
        df_results.to_csv(f, index=False, quoting=csv.QUOTE_ALL, na_rep="NaN")
        # Empty line
        f.write("\n")
        # Observated Concentration
        f.write('"DataType:","Obs Concentration"\n')
        df_obs_conc.to_csv(f, index=False, quoting=csv.QUOTE_ALL, na_rep="NaN")
        # Empty line
        f.write("\n")
